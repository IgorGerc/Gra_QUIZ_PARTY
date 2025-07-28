"""
Moduł logiki gry - zarządza pytaniami, odpowiedziami i punktacją
"""

import openpyxl
import random
from typing import Dict, List, Tuple, Optional, Any
import os
import logging

logger = logging.getLogger(__name__)

class GameLogic:
    """Zarządza logiką gry quiz"""
    
    def __init__(self, questions_file: str = "questions.xlsx"):
        self.questions_file = questions_file
        self.questions: List[Tuple[str, str]] = []  # (pytanie, odpowiedź)
        self.current_question_index = 0
        self.player_scores: Dict[str, int] = {}
        self.current_answers: Dict[str, str] = {}  # gracz -> odpowiedź
        self.current_votes: Dict[str, str] = {}  # gracz -> na co głosuje
        self.correct_answer = ""
        self.game_phase = "waiting"  # waiting, answering, voting, results
        
        self.load_questions()
    
    def load_questions(self):
        """Ładuje pytania z pliku Excel"""
        try:
            if not os.path.exists(self.questions_file):
                self.create_default_questions()
            
            workbook = openpyxl.load_workbook(self.questions_file)
            sheet = workbook.active
            
            self.questions = []
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Pomijamy nagłówek
                if row[0] and row[1]:  # Sprawdź czy pytanie i odpowiedź nie są puste
                    question = str(row[0]).strip()
                    answer = str(row[1]).strip().lower()  # Normalizuj odpowiedź
                    self.questions.append((question, answer))
            
            # Wymieszaj pytania
            random.shuffle(self.questions)
            
            # Ogranicz do 20-30 pytań
            self.questions = self.questions[:25]
            
            logger.info(f"Załadowano {len(self.questions)} pytań")
            
        except Exception as e:
            logger.error(f"Błąd ładowania pytań: {e}")
            self.create_default_questions()
    
    def create_default_questions(self):
        """Tworzy domyślny plik z pytaniami"""
        default_questions = [
            ("Jaka jest stolica Polski?", "warszawa"),
            ("Ile kontynentów jest na Ziemi?", "7"),
            ("Kto napisał 'Pan Tadeusz'?", "adam mickiewicz"),
            ("Jaka jest największa planeta w Układzie Słonecznym?", "jowisz"),
            ("W którym roku Polska wstąpiła do Unii Europejskiej?", "2004"),
            ("Jaki jest najwyższy szczyt w Polsce?", "rysy"),
            ("Ile stron ma trójkąt?", "3"),
            ("Jaka jest waluta w Japonii?", "jen"),
            ("Kto namalował 'Mona Lisę'?", "leonardo da vinci"),
            ("Jaka jest najdłuższa rzeka na świecie?", "nil"),
            ("Ile kości ma dorosły człowiek?", "206"),
            ("W którym roku rozpoczęła się II wojna światowa?", "1939"),
            ("Jaka jest stolica Francji?", "paryż"),
            ("Ile wynosi pierwiastek z 64?", "8"),
            ("Kto wynalazł żarówkę?", "thomas edison"),
            ("Jaka jest najszybsza zwierzę na świecie?", "gepard"),
            ("Ile planet jest w Układzie Słonecznym?", "8"),
            ("Jaka jest najwyższa góra na świecie?", "mount everest"),
            ("W którym roku człowiek po raz pierwszy wylądował na Księżycu?", "1969"),
            ("Jaka jest stolica Włoch?", "rzym"),
            ("Ile wynosi 15% z 200?", "30"),
            ("Kto napisał 'Romeo i Julia'?", "william shakespeare"),
            ("Jaka jest najgłębsza część oceanu?", "rów mariański"),
            ("Ile sekund ma minuta?", "60"),
            ("Jaka jest stolica Hiszpanii?", "madryt"),
            ("Kto skomponował 'Dla Elizy'?", "ludwig van beethoven"),
            ("Jaka jest najdłuższa rzeka w Polsce?", "wisła"),
            ("Ile wynosi 12 x 12?", "144"),
            ("W którym roku upadł Mur Berliński?", "1989"),
            ("Jaka jest stolica Australii?", "canberra")
        ]
        
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Pytania"
            
            # Nagłówki
            sheet['A1'] = "Pytanie"
            sheet['B1'] = "Odpowiedź"
            
            # Dodaj pytania
            for i, (question, answer) in enumerate(default_questions, 2):
                sheet[f'A{i}'] = question
                sheet[f'B{i}'] = answer
            
            workbook.save(self.questions_file)
            self.questions = default_questions
            logger.info(f"Utworzono domyślny plik z {len(default_questions)} pytaniami")
            
        except Exception as e:
            logger.error(f"Błąd tworzenia domyślnych pytań: {e}")
            self.questions = default_questions[:10]  # Fallback
    
    def start_new_game(self):
        """Rozpoczyna nową grę"""
        self.current_question_index = 0
        self.player_scores = {}
        self.current_answers = {}
        self.current_votes = {}
        self.game_phase = "answering"
        logger.info("Rozpoczęto nową grę")
    
    def get_current_question(self) -> Optional[str]:
        """Zwraca aktualne pytanie"""
        if self.current_question_index < len(self.questions):
            question, answer = self.questions[self.current_question_index]
            self.correct_answer = answer
            return question
        return None
    
    def add_player_answer(self, player_name: str, answer: str):
        """Dodaje odpowiedź gracza"""
        if self.game_phase == "answering":
            self.current_answers[player_name] = answer.strip().lower()
            logger.info(f"Gracz {player_name} odpowiedział: {answer}")
    
    def are_all_answers_submitted(self, players_list: List[str]) -> bool:
        """Sprawdza czy wszyscy gracze udzielili odpowiedzi"""
        return len(self.current_answers) >= len(players_list)
    
    def get_grouped_answers(self) -> List[Dict[str, Any]]:
        """Grupuje identyczne odpowiedzi"""
        answer_groups = {}
        
        for player, answer in self.current_answers.items():
            if answer in answer_groups:
                answer_groups[answer]['players'].append(player)
            else:
                answer_groups[answer] = {
                    'answer': answer,
                    'players': [player],
                    'is_correct': self.is_answer_correct(answer)
                }
        
        # Konwertuj na listę i posortuj (poprawne odpowiedzi na górze)
        grouped = list(answer_groups.values())
        grouped.sort(key=lambda x: (not x['is_correct'], x['answer']))
        
        # Sformatuj nazwy graczy
        for group in grouped:
            if len(group['players']) == 1:
                group['players'] = group['players'][0]
            else:
                group['players'] = ", ".join(group['players'])
        
        return grouped
    
    def is_answer_correct(self, answer: str) -> bool:
        """Sprawdza czy odpowiedź jest poprawna"""
        return answer.lower().strip() == self.correct_answer.lower().strip()
    
    def get_players_who_answered_correctly(self) -> List[str]:
        """Zwraca listę graczy, którzy odpowiedzieli poprawnie"""
        correct_players = []
        for player, answer in self.current_answers.items():
            if self.is_answer_correct(answer):
                correct_players.append(player)
        return correct_players
    
    def can_player_vote(self, player_name: str) -> bool:
        """Sprawdza czy gracz może głosować (nie odpowiedział poprawnie)"""
        if player_name not in self.current_answers:
            return False
        return not self.is_answer_correct(self.current_answers[player_name])
    
    def add_vote(self, player_name: str, voted_answer: str):
        """Dodaje głos gracza"""
        if self.game_phase == "voting" and self.can_player_vote(player_name):
            self.current_votes[player_name] = voted_answer.strip().lower()
            logger.info(f"Gracz {player_name} zagłosował na: {voted_answer}")
    
    def are_all_votes_submitted(self, players_list: List[str]) -> bool:
        """Sprawdza czy wszyscy uprawnieni gracze zagłosowali"""
        eligible_voters = [p for p in players_list if self.can_player_vote(p)]
        return len(self.current_votes) >= len(eligible_voters)
    
    def calculate_round_scores(self, players_list: List[str]) -> Dict[str, int]:
        """Oblicza punkty za rundę"""
        round_scores = {player: 0 for player in players_list}
        
        # Inicjalizuj wyniki graczy jeśli to pierwsza runda
        for player in players_list:
            if player not in self.player_scores:
                self.player_scores[player] = 0
        
        # 3 punkty za poprawną odpowiedź od razu
        correct_players = self.get_players_who_answered_correctly()
        for player in correct_players:
            round_scores[player] += 3
            self.player_scores[player] += 3
            logger.info(f"Gracz {player} dostaje 3 pkt za poprawną odpowiedź")
        
        # 2 punkty za zagłosowanie na poprawną odpowiedź (tylko dla tych co nie odpowiedzieli poprawnie)
        for voter, voted_answer in self.current_votes.items():
            if self.is_answer_correct(voted_answer) and voter not in correct_players:
                round_scores[voter] += 2
                self.player_scores[voter] += 2
                logger.info(f"Gracz {voter} dostaje 2 pkt za głos na poprawną odpowiedź")
        
        # 1 punkt za każdy głos na swoją błędną odpowiedź
        for voter, voted_answer in self.current_votes.items():
            for player, player_answer in self.current_answers.items():
                if (player_answer == voted_answer and 
                    not self.is_answer_correct(player_answer)):
                    round_scores[player] += 1
                    self.player_scores[player] += 1
                    logger.info(f"Gracz {player} dostaje 1 pkt za głos na swoją błędną odpowiedź")
        
        return round_scores
    
    def next_question(self):
        """Przechodzi do następnego pytania"""
        self.current_question_index += 1
        self.current_answers = {}
        self.current_votes = {}
        self.game_phase = "answering"
    
    def is_game_finished(self) -> bool:
        """Sprawdza czy gra się skończyła"""
        return self.current_question_index >= len(self.questions)
    
    def get_final_scores(self) -> Dict[str, int]:
        """Zwraca końcowe wyniki"""
        return self.player_scores.copy()
    
    def get_current_scores(self) -> Dict[str, int]:
        """Zwraca aktualne wyniki"""
        return self.player_scores.copy()
    
    def reset_game(self):
        """Resetuje grę"""
        self.current_question_index = 0
        self.player_scores = {}
        self.current_answers = {}
        self.current_votes = {}
        self.game_phase = "waiting"
        
        # Wymieszaj pytania ponownie
        random.shuffle(self.questions)
    
    def get_game_progress(self) -> Tuple[int, int]:
        """Zwraca postęp gry (aktualne pytanie, łączna liczba pytań)"""
        return (self.current_question_index + 1, len(self.questions))
    
    def get_correct_answer(self) -> str:
        """Zwraca poprawną odpowiedź na aktualne pytanie"""
        return self.correct_answer